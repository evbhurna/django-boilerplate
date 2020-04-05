from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.shortcuts import redirect, render
from django.utils.timezone import now
from django.contrib.auth.models import User
from decimal import Decimal
from company.models import Employee
from .models import Attendance
import openpyxl, datetime


today = now

def getTimeDelta(time):
    h = int(time[:2])*3600
    m = int(time[3:5])*60
    s = int(time[6:8])

    timedelta = datetime.timedelta(seconds=h+m+s)
    return timedelta



@login_required
def attendance(request):
    user = Employee.objects.get(user=request.user)
    if request.method == 'POST':
        attendance_file = request.FILES['attendance_file']
        wb = openpyxl.load_workbook(attendance_file)
        worksheet = wb["Sheet1"]
        excel_data = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)

        from django.utils.dateparse import parse_datetime
        import datetime

        for i in range(len(excel_data)-1):
            j = i + 1

            employee = Employee.objects.get(
                employee_number = excel_data[j][1],
                last_name__iexact = excel_data[j][0]
            )

            in_date = parse_datetime(excel_data[j][2])
            time_in = in_date + getTimeDelta(excel_data[j][3])

            out_date = parse_datetime(excel_data[j][4])
            time_out = out_date + getTimeDelta(excel_data[j][5])

            check = Attendance.objects.filter(
                time_in__day = in_date.day,
                time_out__day = out_date.day
            )

            if check:
                pass
            else:
                attendance_created = Attendance.objects.create(
                    employee = employee,
                    time_in = time_in,
                    time_out = time_out,
                    type = excel_data[j][6]
                )

        messages.success(
            request, 'Attendance file successfully uploaded.')
        return redirect('/attendance/')
    else:
        attendance = Attendance.objects.filter(employee__company=user.company)
        return render(request, 'attendance/attendance.html', {
            'user':user, 'attendance': attendance, 'today':today
        })

@login_required
def deductionsUpload(request):
    user = Employee.objects.get(user=request.user)
    if request.FILES['deduction_file']:
        deduction_file = request.FILES['deduction_file']
        wb = openpyxl.load_workbook(deduction_file)
        worksheet = wb["Sheet1"]
        excel_data = list()
        for row in worksheet.iter_rows():
            row_data = list()
            for cell in row:
                row_data.append(str(cell.value))
            excel_data.append(row_data)
        from django.utils.dateparse import parse_datetime
        for i in range(len(excel_data)-1):
            j = i + 1
            employee = Employee.objects.get(
                employee_number = excel_data[j][1],
                last_name__iexact = excel_data[j][0]
            ) 
            deduction = Deduction.objects.create(
                employee = employee,
                name = excel_data[j][2],
                amount = excel_data[j][5],
                date = parse_datetime(excel_data[j][4]).date(),
                notes = excel_data[j][3]
            )
        
        messages.success(
            request, 'Deduction/s file successfully uploaded.')
        return redirect('/deductions/')
    else:
        pass